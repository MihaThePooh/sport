#!/usr/bin/env python3.12
"""Собирает из data/log.csv читаемую книгу Excel.

Запускать python3.12 — openpyxl стоит только там (python3 на этом VPS
десятый и без него).

Таблица широкая: строка — занятие, колонки — упражнения. Так видно
и подходы, и сумму, и чем этот день отличался от соседних.

Зелёная заливка ставится сама: раньше Михаил красил рекордные ячейки
руками, а рекорд — это просто максимум по колонке на текущий момент.
Считаем нарастающим итогом, то есть красим тот день, когда рекорд был
поставлен, а не все, что его позже повторили.
"""
import csv
import os
import sys
import tempfile
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent
# Ту же переменную читает server.py: тесты и боевой дневник не пересекаются.
DATA = Path(os.environ.get("SPORT_DATA") or (BASE / "data"))
LOG = DATA / "log.csv"
OUT = DATA / "тренировки.xlsx"

MOVES = ["подтягивания", "отжимания", "приседания"]

GREEN = PatternFill("solid", fgColor="C6E7D2")   # рекорд за 12 месяцев
GOLD = PatternFill("solid", fgColor="F3E3B0")    # рекорд за всё время
YEAR = 365
HEADFILL = PatternFill("solid", fgColor="2F6F4E")
STRIPE = PatternFill("solid", fgColor="F7F6F4")  # через строку, чтобы глаз не сползал
THIN = Side(style="thin", color="E3DED7")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load():
    """CSV → {дата: {упражнение: [подходы]}}, отсортировано по дате."""
    days = {}
    if not LOG.exists():
        return days
    with LOG.open(encoding="utf-8", newline="") as f:
        for r in csv.DictReader(f):
            when = (r.get("дата") or "").strip()
            move = (r.get("упражнение") or "").strip()
            if not when or move not in MOVES:
                continue
            sets = [int(x) for x in str(r.get("подходы", "")).split() if x.isdigit()]
            total = str(r.get("всего", "")).strip()
            total = int(total) if total.isdigit() else sum(sets)
            # Занятие без подходов, но с суммой, — тоже занятие: в старой
            # таблице есть дни, где записана только она.
            if sets or total:
                days.setdefault(when, {})[move] = {"подходы": sets, "всего": total}
    return dict(sorted(days.items()))


def build():
    days = load()
    wb = Workbook()
    ws = wb.active
    ws.title = "Тренировки"

    # Шапка в два яруса: над каждым упражнением — «подходы» и «всего».
    ws.cell(1, 1, "дата")
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    col = 2
    where = {}
    for m in MOVES:
        ws.cell(1, col, m)
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        ws.cell(2, col, "подходы")
        ws.cell(2, col + 1, "всего")
        where[m] = col
        col += 2

    for row in (1, 2):
        for c in range(1, col):
            cell = ws.cell(row, c)
            cell.fill = HEADFILL
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BOX

    # Рекорды копим по ходу: красится день прорыва, а не его повторы.
    #
    # Два вида отметок, и это не украшательство. Абсолютные рекорды
    # Михаил поставил осенью 2022-го, когда делал по 7 подходов за раз.
    # Сейчас подходов 3 — сумма за занятие физически не догонит ту,
    # и зелёного в таблице не было бы с 2022 года. Поэтому основная,
    # живая отметка — рекорд за последние 12 месяцев на момент занятия:
    # он честно говорит «лучше, чем ты был весь этот год». Абсолютный
    # пик отмечается отдельно, золотым, как памятка.
    peak = {m: {"подход": 0, "всего": 0} for m in MOVES}
    r = 3

    def year_best(m, upto):
        """Лучшее за 12 месяцев ДО этого дня (сам день не считаем)."""
        edge = upto - timedelta(days=YEAR)
        rep = tot = 0
        for d, mv in days.items():
            day = date.fromisoformat(d)
            if not (edge <= day < upto) or m not in mv:
                continue
            if mv[m]["подходы"]:
                rep = max(rep, max(mv[m]["подходы"]))
            tot = max(tot, mv[m]["всего"])
        return rep, tot

    for when, moves in days.items():
        d = date.fromisoformat(when)
        cell = ws.cell(r, 1, d)
        cell.number_format = "DD.MM.YYYY"
        cell.alignment = Alignment(horizontal="center")

        for m in MOVES:
            c = where[m]
            got = moves.get(m)
            if not got:
                for k in (c, c + 1):
                    ws.cell(r, k, "").alignment = Alignment(horizontal="center")
                continue
            sets, total = got["подходы"], got["всего"]
            a = ws.cell(r, c, " + ".join(str(x) for x in sets) if sets else "—")
            b = ws.cell(r, c + 1, total)
            a.alignment = Alignment(horizontal="center")
            b.alignment = Alignment(horizontal="center")
            b.font = Font(bold=True)

            yrep, ytot = year_best(m, d)
            if sets and max(sets) > peak[m]["подход"]:
                a.fill = GOLD                      # рекорд за всё время
                peak[m]["подход"] = max(sets)
            elif sets and yrep and max(sets) > yrep:
                a.fill = GREEN                     # лучше, чем весь прошедший год
            if total > peak[m]["всего"]:
                b.fill = GOLD
                peak[m]["всего"] = total
            elif ytot and total > ytot:
                b.fill = GREEN

        if r % 2:
            for c in range(1, col):
                if not ws.cell(r, c).fill.fgColor.rgb or \
                        ws.cell(r, c).fill.fill_type is None:
                    ws.cell(r, c).fill = STRIPE
        for c in range(1, col):
            ws.cell(r, c).border = BOX
        r += 1

    # Итоговая строка: сколько всего сделано за всё время.
    ws.cell(r, 1, "итого").font = Font(bold=True)
    for m in MOVES:
        c = where[m]
        s = sum(v[m]["всего"] for v in days.values() if m in v)
        n = sum(1 for v in days.values() if m in v)
        ws.cell(r, c, "%d занятий" % n).alignment = Alignment(horizontal="center")
        cell = ws.cell(r, c + 1, s)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for c in range(1, col):
        ws.cell(r, c).border = BOX

    ws.column_dimensions["A"].width = 13
    for m in MOVES:
        ws.column_dimensions[get_column_letter(where[m])].width = 16
        ws.column_dimensions[get_column_letter(where[m] + 1)].width = 8
    ws.freeze_panes = "B3"          # шапка и дата не уезжают при прокрутке

    OUT.parent.mkdir(parents=True, exist_ok=True)
    # Сохраняем через временный файл и подменяем одним движением.
    # Книга пересобирается после каждой записи, и без этого можно было
    # бы скачать её ровно в тот момент, когда openpyxl дописывает архив,
    # — получился бы битый .xlsx без единой ошибки на экране.
    fd, tmp = tempfile.mkstemp(dir=str(OUT.parent), suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        # mkstemp отдаёт 0600 — своим же файлом потом не поделишься.
        os.chmod(tmp, 0o644)
        os.replace(tmp, OUT)
    except BaseException:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise
    return len(days)


if __name__ == "__main__":
    n = build()
    print("собрано занятий: %d → %s" % (n, OUT))
    if not n:
        sys.exit(1)
