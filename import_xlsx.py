#!/usr/bin/env python3.12
"""Переносит старую таблицу тренировок в data/log.csv.

Запуск: python3.12 import_xlsx.py <файл.xlsx> [--обновить]
Без --обновить только показывает, что получится, и ничего не пишет.

Старая таблица накопилась за три года и по дороге обросла тремя
болячками — их и лечим:

1. Даты в двух видах: настоящая дата Excel и текст «28,10,2024».
2. Excel съел часть подходов. Ячейку «12-12-15» он счёл датой
   15 декабря 2012 года. Разбираем обратно: год берём по двум
   последним цифрам, дальше месяц и день. Проверка железная —
   восстановленная сумма обязана совпасть с колонкой «сумма»,
   которую Excel не трогал. Не совпала — не восстанавливаем.
3. Приседания записаны только подходами, без колонки суммы,
   и иногда одним числом вместо списка.

Колонки исходника:
    A дата | B подтягивания сумма | C подходы
           | D отжимания сумма    | E подходы
           | F приседания — сразу подходы
"""
import argparse
import csv
import datetime
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
OUT = BASE / "data" / "log.csv"
HEAD = ["дата", "упражнение", "подходы", "всего"]

# упражнение -> (колонка суммы или None, колонка подходов)
COLS = {
    "подтягивания": (2, 3),
    "отжимания": (4, 5),
    "приседания": (None, 6),
}


def parse_day(v):
    """Дата из ячейки: настоящая дата Excel либо текст «ДД,ММ,ГГГГ»."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v or "").strip()
    m = re.fullmatch(r"(\d{1,2})\s*[.,/-]\s*(\d{1,2})\s*[.,/-]\s*(\d{2,4})", s)
    if not m:
        return None
    d, mo, y = (int(x) for x in m.groups())
    if y < 100:
        y += 2000
    try:
        return datetime.date(y, mo, d)
    except ValueError:
        return None


def parse_sets(v, total, notes, where):
    """Подходы из ячейки. Возвращает список чисел."""
    if v is None or v == "":
        return []
    if isinstance(v, int):
        return [v]
    if isinstance(v, (datetime.datetime, datetime.date)):
        # Ячейку испортил Excel. Восстанавливаем и сверяем с суммой.
        guess = [v.year % 100, v.month, v.day]
        if isinstance(total, int) and sum(guess) == total:
            notes.append("%s: восстановлено из даты %s → %s" %
                         (where, v.strftime("%Y-%m-%d"), guess))
            return guess
        notes.append("%s: ячейка испорчена Excel (%s), сумма не сошлась — "
                     "оставил пустой" % (where, v))
        return []
    nums = [int(x) for x in re.findall(r"\d+", str(v))]
    return [n for n in nums if 0 < n <= 500]


def convert(path):
    ws = load_workbook(path, data_only=True).worksheets[0]
    rows, notes, skipped = [], [], []

    for r in range(2, ws.max_row + 1):
        day = parse_day(ws.cell(r, 1).value)
        if not day:
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, 7)):
                skipped.append("строка %d: не разобрал дату %r" %
                               (r, ws.cell(r, 1).value))
            continue

        for move, (tcol, scol) in COLS.items():
            total = ws.cell(r, tcol).value if tcol else None
            where = "строка %d, %s" % (r, move)
            sets = parse_sets(ws.cell(r, scol).value, total, notes, where)

            if not sets:
                # Сумма есть, а подходов нет — занятие было, детали утеряны.
                if isinstance(total, int) and total > 0:
                    rows.append({"дата": day.isoformat(), "упражнение": move,
                                 "подходы": "", "всего": total})
                    notes.append("%s: подходы не записаны, сохранил только "
                                 "сумму %d" % (where, total))
                continue

            if isinstance(total, int) and sum(sets) != total:
                notes.append("%s: подходы %s = %d, а в таблице %d — "
                             "оставил подходы" %
                             (where, " ".join(map(str, sets)), sum(sets), total))
            rows.append({"дата": day.isoformat(), "упражнение": move,
                         "подходы": " ".join(map(str, sets)),
                         "всего": sum(sets)})

    rows.sort(key=lambda x: (x["дата"], list(COLS).index(x["упражнение"])))
    return rows, notes, skipped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--обновить", action="store_true",
                    help="записать data/log.csv (без флага только показывает)")
    a = ap.parse_args()

    rows, notes, skipped = convert(a.xlsx)
    if not rows:
        print("ничего не разобрал", file=sys.stderr)
        return 1

    by = {}
    for x in rows:
        by[x["упражнение"]] = by.get(x["упражнение"], 0) + 1
    print("занятий всего: %d, с %s по %s" %
          (len(rows), rows[0]["дата"], rows[-1]["дата"]))
    for m, n in by.items():
        print("  %-14s %3d записей" % (m, n))

    if notes:
        print("\nчто пришлось поправить (%d):" % len(notes))
        for n in notes:
            print("  " + n)
    if skipped:
        print("\nпропущено (%d):" % len(skipped))
        for s in skipped:
            print("  " + s)

    if not a.обновить:
        print("\nничего не записал. Флаг --обновить запишет в %s" % OUT)
        return 0

    if OUT.exists():
        keep = OUT.with_suffix(".csv.бэкап-%s" %
                               datetime.date.today().isoformat())
        keep.write_bytes(OUT.read_bytes())
        print("\nстарый log.csv сохранён в %s" % keep.name)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=HEAD, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print("записано: %s" % OUT)
    return 0


if __name__ == "__main__":
    sys.exit(main())
